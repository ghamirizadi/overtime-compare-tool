import tkinter as tk
from tkinter import filedialog, messagebox
from pathlib import Path
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

KEY_COL = "شماره پرسنلی"
OT_COL = "اضافه کار"

# رنگ‌ها
FILL_OK = PatternFill("solid", fgColor="D9E1F2")     # آبی کم‌رنگ
FILL_DIFF = PatternFill("solid", fgColor="F8CBAD")   # قرمز کم‌رنگ
FILL_MISS = PatternFill("solid", fgColor="FFF2CC")   # زرد کم‌رنگ

HEADER_FILL = PatternFill("solid", fgColor="E7E6E6")
HEADER_FONT = Font(bold=True)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)


def _require_columns(df: pd.DataFrame, file_label: str):
    cols = list(df.columns)
    if KEY_COL not in cols or OT_COL not in cols:
        raise ValueError(
            f"در {file_label} ستون‌های لازم پیدا نشد.\n"
            f"باید دقیقاً ستون‌های «{KEY_COL}» و «{OT_COL}» وجود داشته باشند.\n"
            f"ستون‌های موجود: {cols}"
        )


def _prep_df(path: str, prefix: str):
    df = pd.read_excel(path)
    _require_columns(df, prefix)

    # نرمال‌سازی کلید: رشته، حفظ صفرهای اول، حذف فاصله
    df[KEY_COL] = df[KEY_COL].astype(str).str.strip()

    # نرمال‌سازی اضافه‌کار: عدد صحیح
    df[OT_COL] = pd.to_numeric(df[OT_COL], errors="coerce").fillna(0).astype("int64")

    # کنترل تکراری‌ها (اگر باشد، برای مقایسه OT را جمع می‌زنیم و بقیه ستون‌ها را از اولین ردیف نگه می‌داریم)
    dup_keys = df[df[KEY_COL].duplicated(keep=False)][KEY_COL].unique().tolist()
    if dup_keys:
        other_cols = [c for c in df.columns if c not in (KEY_COL, OT_COL)]
        agg = {OT_COL: "sum"}
        for c in other_cols:
            agg[c] = "first"
        df = df.groupby(KEY_COL, as_index=False).agg(agg)
    else:
        dup_keys = []

    # نام‌گذاری ستون‌ها برای جلوگیری از تداخل
    rename_map = {}
    for c in df.columns:
        if c == KEY_COL:
            continue
        rename_map[c] = f"{prefix}_{c}"
    df = df.rename(columns=rename_map)

    return df, dup_keys


def build_report(file_a: str, file_b: str, out_path: str):
    a, dup_a = _prep_df(file_a, "A")
    b, dup_b = _prep_df(file_b, "B")

    merged = a.merge(b, on=KEY_COL, how="outer")

    # ستون‌های مقایسه
    ot_a = f"A_{OT_COL}"
    ot_b = f"B_{OT_COL}"

    # ممکنه در یک سمت NaN باشد
    merged[ot_a] = merged.get(ot_a)
    merged[ot_b] = merged.get(ot_b)

    def status_row(r):
        a_na = pd.isna(r.get(ot_a))
        b_na = pd.isna(r.get(ot_b))
        if a_na and not b_na:
            return "MISSING_IN_A"
        if b_na and not a_na:
            return "MISSING_IN_B"
        if a_na and b_na:
            return "MISSING_BOTH"
        return "OK" if int(r[ot_a]) == int(r[ot_b]) else "DIFF"

    merged["Status"] = merged.apply(status_row, axis=1)
    merged["Delta"] = merged.apply(
        lambda r: (int(r[ot_a]) - int(r[ot_b]))
        if (pd.notna(r.get(ot_a)) and pd.notna(r.get(ot_b))) else pd.NA,
        axis=1
    )

    # ترتیب ستون‌ها: کلید، OTها، Delta/Status، سپس بقیه
    cols = list(merged.columns)
    front = [KEY_COL, ot_a, ot_b, "Delta", "Status"]
    rest = [c for c in cols if c not in front]
    merged = merged[front + rest].sort_values(["Status", KEY_COL], ascending=[True, True])

    summary = merged["Status"].value_counts().rename_axis("Status").reset_index(name="Count")

    diffs = merged[merged["Status"] != "OK"].copy()

    # خروجی اولیه با pandas
    with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
        merged.to_excel(writer, sheet_name="Comparison", index=False)
        diffs.to_excel(writer, sheet_name="Diffs_Only", index=False)
        summary.to_excel(writer, sheet_name="Summary", index=False)

        # اگر تکراری داشتیم
        if dup_a or dup_b:
            pd.DataFrame({
                "Duplicates_in_A": dup_a + [pd.NA] * max(0, len(dup_b) - len(dup_a)),
                "Duplicates_in_B": dup_b + [pd.NA] * max(0, len(dup_a) - len(dup_b)),
            }).to_excel(writer, sheet_name="Duplicates", index=False)

    # رنگ‌آمیزی با openpyxl
    wb = load_workbook(out_path)
    ws = wb["Comparison"]

    # استایل هدر
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER

    # پیدا کردن ستون Status
    header = [c.value for c in ws[1]]
    try:
        status_col_idx = header.index("Status") + 1
    except ValueError:
        status_col_idx = None

    # رنگ ردیف‌ها بر اساس Status
    for row in range(2, ws.max_row + 1):
        status_val = ws.cell(row=row, column=status_col_idx).value if status_col_idx else None

        if status_val == "OK":
            fill = FILL_OK
        elif status_val == "DIFF":
            fill = FILL_DIFF
        elif status_val in ("MISSING_IN_A", "MISSING_IN_B", "MISSING_BOTH"):
            fill = FILL_MISS
        else:
            fill = None

        if fill:
            for col in range(1, ws.max_column + 1):
                ws.cell(row=row, column=col).fill = fill

    # تنظیم عرض ستون‌ها (ساده و کاربردی)
    for col in range(1, ws.max_column + 1):
        letter = get_column_letter(col)
        ws.column_dimensions[letter].width = 18

    wb.save(out_path)


# ---------------- GUI ----------------
class App(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("مقایسه اضافه‌کار دو فایل اکسل")
        self.geometry("720x260")
        self.resizable(False, False)

        self.file_a_var = tk.StringVar()
        self.file_b_var = tk.StringVar()
        self.out_var = tk.StringVar(value=str(Path.cwd() / "report.xlsx"))

        self._build()

    def _build(self):
        pad = {"padx": 10, "pady": 8}

        tk.Label(self, text="فایل A:").grid(row=0, column=0, sticky="w", **pad)
        tk.Entry(self, textvariable=self.file_a_var, width=70).grid(row=0, column=1, **pad)
        tk.Button(self, text="Browse", command=self.pick_a).grid(row=0, column=2, **pad)

        tk.Label(self, text="فایل B:").grid(row=1, column=0, sticky="w", **pad)
        tk.Entry(self, textvariable=self.file_b_var, width=70).grid(row=1, column=1, **pad)
        tk.Button(self, text="Browse", command=self.pick_b).grid(row=1, column=2, **pad)

        tk.Label(self, text="مسیر خروجی:").grid(row=2, column=0, sticky="w", **pad)
        tk.Entry(self, textvariable=self.out_var, width=70).grid(row=2, column=1, **pad)
        tk.Button(self, text="Save As", command=self.pick_out).grid(row=2, column=2, **pad)

        tk.Button(self, text="ساخت گزارش", height=2, command=self.run).grid(row=3, column=1, sticky="ew", padx=10, pady=18)

        tk.Label(
            self,
            text=f"ستون‌های ضروری: «{KEY_COL}» و «{OT_COL}» در هر دو فایل باید وجود داشته باشند.",
        ).grid(row=4, column=0, columnspan=3, sticky="w", padx=10)

    def pick_a(self):
        p = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx *.xls")])
        if p:
            self.file_a_var.set(p)

    def pick_b(self):
        p = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx *.xls")])
        if p:
            self.file_b_var.set(p)

    def pick_out(self):
        p = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel files", "*.xlsx")])
        if p:
            self.out_var.set(p)

    def run(self):
        fa = self.file_a_var.get().strip()
        fb = self.file_b_var.get().strip()
        out = self.out_var.get().strip()

        if not fa or not fb:
            messagebox.showerror("خطا", "هر دو فایل A و B را انتخاب کن.")
            return

        try:
            build_report(fa, fb, out)
            messagebox.showinfo("انجام شد", f"گزارش ساخته شد:\n{out}")
        except Exception as e:
            messagebox.showerror("خطا", str(e))


if __name__ == "__main__":
    App().mainloop()